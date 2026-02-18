
import asyncio
import io
import csv
import openpyxl
from unittest.mock import AsyncMock, MagicMock
from app.api.endpoints.sensor import export_sensor_data
from app.models.processed_sensor_data import ProcessedSensorData
from datetime import datetime

async def verify_export():
    print("Verifying Export Feature...")
    
    # Mock DB
    db = AsyncMock()
    mock_result = MagicMock()
    
    # Create sample data
    sample_data = [
        ProcessedSensorData(
            id=1,
            device_id=1,
            created_at=datetime.now(),
            status="SAFE",
            tilt_diff_x=0.1, tilt_diff_y=0.2, tilt_diff_z=0.3,
            distance_diff_mm=5.0,
            tilt_change_percent=1.0, distance_change_percent=0.5
        ),
        ProcessedSensorData(
            id=2,
            device_id=1,
            created_at=datetime.now(),
            status="ALERT",
            tilt_diff_x=10.0, tilt_diff_y=0.0, tilt_diff_z=0.0,
            distance_diff_mm=-2.0,
            tilt_change_percent=20.0, distance_change_percent=0.2
        )
    ]
    
    mock_result.scalars.return_value.all.return_value = sample_data
    db.execute.return_value = mock_result
    
    # Test CSV Export
    print("Testing CSV Export...")
    response_csv = await export_sensor_data(device_id=1, format="csv", db=db)
    
    # Read response content
    content_csv = b""
    async for chunk in response_csv.body_iterator:
        content_csv += chunk
        
    text_csv = content_csv.decode()
    print(f"CSV Content (first 100 chars): {text_csv[:100]}...")
    
    # Validate CSV validity
    reader = csv.reader(io.StringIO(text_csv))
    rows = list(reader)
    if len(rows) == 3: # Header + 2 rows
        print("✅ CSV Header and Row count correct")
    else:
        print(f"❌ CSV Row count incorrect: {len(rows)}")

    # Test Excel Export
    print("\nTesting Excel Export...")
    response_xlsx = await export_sensor_data(device_id=1, format="xlsx", db=db)
    
    content_xlsx = b""
    async for chunk in response_xlsx.body_iterator:
        content_xlsx += chunk
        
    # Validate Excel validity
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content_xlsx))
        sheet = wb.active
        if sheet.max_row == 3:
             print("✅ Excel Row count correct")
        else:
             print(f"❌ Excel Row count incorrect: {sheet.max_row}")
    except Exception as e:
        print(f"❌ Excel invalid: {e}")

if __name__ == "__main__":
    asyncio.run(verify_export())
