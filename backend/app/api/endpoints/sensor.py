from fastapi import APIRouter, Depends, HTTPException, status, UploadFile, File
import os
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
from typing import List, Literal
import csv
import io
import asyncio
from datetime import datetime, timezone
from dateutil.parser import parse as parse_date
import pytz
import openpyxl
from fastapi.responses import StreamingResponse

from app.core.db import get_db
from app.schemas.sensor import SensorIngestRequest, ProcessedSensorDataResponse, ManualSensorIngestRequest
from app.models.processed_sensor_data import ProcessedSensorData
from app.services.device_service import DeviceService
from app.services.sensor_service import SensorService

router = APIRouter()

@router.post("/ingest", response_model=ProcessedSensorDataResponse, status_code=status.HTTP_201_CREATED)
async def ingest_sensor_data(
    sensor_data: SensorIngestRequest,
    db: AsyncSession = Depends(get_db)
):
    """
    Ingest sensor data from ESP32 device.
    """
    # Validate device exists
    device = await DeviceService.get_device_by_uid(db, sensor_data.device_uid)
    
    if not device:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Device with UID '{sensor_data.device_uid}' not found. "
                   f"Device must be registered before sending data."
        )
    
    # Process and store sensor reading
    reading = await SensorService.ingest_sensor_data(db, device, sensor_data)
    
    return reading

@router.post("/ingest/manual", response_model=ProcessedSensorDataResponse, status_code=status.HTTP_201_CREATED)
async def ingest_manual_sensor_data(
    sensor_data: ManualSensorIngestRequest,
    db: AsyncSession = Depends(get_db)
):
    """
    Manually ingest sensor data with optional custom timestamp.
    """
    # Validate device exists
    device = await DeviceService.get_device_by_uid(db, sensor_data.device_uid)
    
    if not device:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Device with UID '{sensor_data.device_uid}' not found."
        )
    
    # Process and store sensor reading with custom timestamp
    reading = await SensorService.ingest_sensor_data(
        db, 
        device, 
        sensor_data, 
        timestamp=sensor_data.timestamp
    )
    
    return reading

@router.post("/devices/{device_id}/upload", status_code=status.HTTP_201_CREATED)
async def upload_processed_data(
    device_id: int,
    file: UploadFile = File(None),
    db: AsyncSession = Depends(get_db)
):
    """
    Upload processed sensor data from a CSV or Excel file and associate it with a device.
    Generates dummy raw sensor data (0s) to satisfy foreign key constraints.
    If no file is provided, attempts to read from a local temporary 'temp_uploads' folder.
    """
    device = await DeviceService.get_device_by_id(db, device_id)
    if not device:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Device with ID '{device_id}' not found."
        )

    # Temporary storage logic
    temp_dir = os.path.join(os.getcwd(), "temp_uploads")
    default_file_path = os.path.join(temp_dir, "sensor_data_processed.xlsx")

    content = None
    filename = ""

    if file:
        content = await file.read()
        filename = file.filename.lower()
        
        # Save to temp storage for future re-use
        os.makedirs(temp_dir, exist_ok=True)
        with open(default_file_path, "wb") as f:
            f.write(content)
            
    else:
        # Try to read from temp storage
        if not os.path.exists(default_file_path):
            raise HTTPException(
                status_code=400, 
                detail="No file provided and no previously uploaded file found in temporary storage."
            )
        with open(default_file_path, "rb") as f:
            content = f.read()
        filename = "sensor_data_processed.xlsx"
    
    rows_processed = 0
    errors = []

    try:
        from app.models.raw_sensor_data import RawSensorData
        from app.models.processed_sensor_data import ProcessedSensorData
        
        if filename.endswith(".csv"):
            import codecs
            reader = csv.DictReader(codecs.iterdecode(io.BytesIO(content), 'utf-8'))
            
            for index, row in enumerate(reader):
                try:
                    # Map columns (fallback to lowercase matching if needed, but exact matches preferred)
                    def get_val(keys, default=0.0):
                        for k in keys:
                            if k in row and row[k]:
                                return float(row[k])
                        return default
                        
                    created_at_raw = row.get("Created At") or row.get("created_at")
                    if created_at_raw:
                        created_at = parse_date(created_at_raw)
                        if created_at.tzinfo is None:
                            # Assume IST for naive uploaded times
                            ist_tz = pytz.timezone('Asia/Kolkata')
                            created_at = ist_tz.localize(created_at).astimezone(pytz.UTC)
                    else:
                        created_at = datetime.now(timezone.utc)

                    status_val = row.get("Status") or row.get("status", "SAFE")
                    tilt_x = get_val(["Tilt Diff X", "tilt_diff_x"])
                    tilt_y = get_val(["Tilt Diff Y", "tilt_diff_y"])
                    tilt_z = get_val(["Tilt Diff Z", "tilt_diff_z"])
                    dist_mm = get_val(["Distance Diff (mm)", "distance_diff_mm"])
                    tilt_pct = get_val(["Tilt Change %", "tilt_change_percent"])
                    dist_pct = get_val(["Distance Change %", "distance_change_percent"])
                    
                    # 1. Create Mock Raw Data
                    mock_raw = RawSensorData(
                        device_id=device_id,
                        tilt_x=0.0, tilt_y=0.0, tilt_z=0.0, distance_mm=0.0,
                        created_at=created_at
                    )
                    db.add(mock_raw)
                    await db.flush() # Flush to get ID
                    
                    # 2. Create Processed Data using real values
                    processed = ProcessedSensorData(
                        device_id=device_id,
                        raw_data_id=mock_raw.id,
                        tilt_diff_x=tilt_x,
                        tilt_diff_y=tilt_y,
                        tilt_diff_z=tilt_z,
                        distance_diff_mm=dist_mm,
                        tilt_change_percent=tilt_pct,
                        distance_change_percent=dist_pct,
                        status=status_val,
                        created_at=created_at
                    )
                    db.add(processed)
                    rows_processed += 1
                except Exception as e:
                    errors.append(f"Row {index+2}: {str(e)}")
                    
        elif filename.endswith(".xlsx"):
            workbook = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
            sheet = workbook.active
            headers = [cell.value for cell in sheet[1]]
            
            for index, row_cells in enumerate(sheet.iter_rows(min_row=2, values_only=True)):
                if not any(row_cells): continue # Skip empty rows
                
                try:
                    row = dict(zip(headers, row_cells))
                    
                    def get_val(keys, default=0.0):
                        for k in keys:
                            if k in row and row[k] is not None:
                                return float(row[k])
                        return default
                    
                    created_at_raw = row.get("Created At") or row.get("created_at")
                    if isinstance(created_at_raw, datetime):
                        created_at = created_at_raw
                    elif created_at_raw:
                        created_at = parse_date(str(created_at_raw))
                    else:
                        created_at = datetime.now(timezone.utc)
                        
                    if created_at.tzinfo is None:
                        # Assume IST for naive uploaded times
                        ist_tz = pytz.timezone('Asia/Kolkata')
                        created_at = ist_tz.localize(created_at).astimezone(pytz.UTC)

                    status_val = str(row.get("Status") or row.get("status") or "SAFE")
                    tilt_x = get_val(["Tilt Diff X", "tilt_diff_x"])
                    tilt_y = get_val(["Tilt Diff Y", "tilt_diff_y"])
                    tilt_z = get_val(["Tilt Diff Z", "tilt_diff_z"])
                    dist_mm = get_val(["Distance Diff (mm)", "distance_diff_mm"])
                    tilt_pct = get_val(["Tilt Change %", "tilt_change_percent"])
                    dist_pct = get_val(["Distance Change %", "distance_change_percent"])
                    
                    # 1. Create Mock Raw Data
                    mock_raw = RawSensorData(
                        device_id=device_id,
                        tilt_x=0.0, tilt_y=0.0, tilt_z=0.0, distance_mm=0.0,
                        created_at=created_at
                    )
                    db.add(mock_raw)
                    await db.flush() # Flush to get ID
                    
                    # 2. Create Processed Data
                    processed = ProcessedSensorData(
                        device_id=device_id,
                        raw_data_id=mock_raw.id,
                        tilt_diff_x=tilt_x,
                        tilt_diff_y=tilt_y,
                        tilt_diff_z=tilt_z,
                        distance_diff_mm=dist_mm,
                        tilt_change_percent=tilt_pct,
                        distance_change_percent=dist_pct,
                        status=status_val,
                        created_at=created_at
                    )
                    db.add(processed)
                    rows_processed += 1
                except Exception as e:
                    errors.append(f"Row {index+2}: {str(e)}")
        else:
            raise HTTPException(status_code=400, detail="Invalid file format. Only .csv and .xlsx are supported.")

        if rows_processed > 0:
            await db.commit()
        else:
            await db.rollback()
            
        return {
            "message": f"Successfully processed {rows_processed} entries.",
            "errors": errors if errors else None
        }
        
    except Exception as e:
        await db.rollback()
        raise HTTPException(status_code=500, detail=f"An error occurred while processing the file: {str(e)}")

@router.get("/devices/{device_id}/processed", response_model=List[ProcessedSensorDataResponse])
async def get_processed_data(
    device_id: int,
    limit: int = 100,
    db: AsyncSession = Depends(get_db)
):
    """
    Get processed sensor data for a specific device.
    """
    result = await db.execute(
        select(ProcessedSensorData)
        .where(ProcessedSensorData.device_id == device_id)
        .order_by(ProcessedSensorData.created_at.desc())
        .limit(limit)
    )
    return result.scalars().all()


@router.get("/devices/{device_id}/export")
async def export_sensor_data(
    device_id: int,
    format: Literal["csv", "xlsx"] = "csv",
    limit: int = 1000,
    db: AsyncSession = Depends(get_db)
):
    """
    Export processed sensor data as CSV or Excel.
    """
    # 1. Fetch Data
    result = await db.execute(
        select(ProcessedSensorData)
        .where(ProcessedSensorData.device_id == device_id)
        .order_by(ProcessedSensorData.created_at.desc())
        .limit(limit)
    )
    data = result.scalars().all()
    
    filename = f"sensor_data_device_{device_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}"

    if format == "csv":
        async def iter_csv():
            output = io.StringIO()
            writer = csv.writer(output)
            
            # Headers
            headers = [
                "ID", "Created At", "Status",
                "Tilt Diff X", "Tilt Diff Y", "Tilt Diff Z",
                "Distance Diff (mm)",
                "Tilt Change %", "Distance Change %"
            ]
            writer.writerow(headers)
            output.seek(0)
            yield output.read()
            output.truncate(0)
            output.seek(0)
            
            # Rows
            for row in data:
                writer.writerow([
                    row.id,
                    row.created_at.isoformat(),
                    row.status,
                    row.tilt_diff_x,
                    row.tilt_diff_y,
                    row.tilt_diff_z,
                    row.distance_diff_mm,
                    row.tilt_change_percent,
                    row.distance_change_percent
                ])
                output.seek(0)
                yield output.read()
                output.truncate(0)
                output.seek(0)

        return StreamingResponse(
            iter_csv(),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename={filename}.csv"}
        )
        
    elif format == "xlsx":
        def generate_excel():
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Sensor Data"
            
            # Headers
            headers = [
                "ID", "Created At", "Status",
                "Tilt Diff X", "Tilt Diff Y", "Tilt Diff Z",
                "Distance Diff (mm)",
                "Tilt Change %", "Distance Change %"
            ]
            sheet.append(headers)
            
            # Rows
            for row in data:
                sheet.append([
                    row.id,
                    row.created_at.replace(tzinfo=None), # Excel doesn't like timezone aware datetimes sometimes
                    row.status,
                    row.tilt_diff_x,
                    row.tilt_diff_y,
                    row.tilt_diff_z,
                    row.distance_diff_mm,
                    row.tilt_change_percent,
                    row.distance_change_percent
                ])
                
            output = io.BytesIO()
            workbook.save(output)
            output.seek(0)
            return output

        loop = asyncio.get_running_loop()
        output = await loop.run_in_executor(None, generate_excel)
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}.xlsx"}
        )
